#! /usr/bin/env python
# -*- coding: utf-8 -*-

"""Parse AIMS TI files to produce valid Catami project
v1.0

expects file structure as

main_excel_file.xlsx
  |
  |--Image Dir 01
  |       |--images01.jpg
  |       |--images02.jpg
  |       |-- ...
  |--Image Dir 02
  |       |--images01.jpg
  |       |--images02.jpg
  |       |-- ...
  ...
Most data is taken from the XLSX file (inspected JPG files have lat/longs in EXIF, XLSX lat/longs are considered authoratative)
Camera details are taken from image EXIF data
Data without source are set to fill (-999 for numerical data, 'null' for string data)

v1.0 28/05/2013 markg@ivec.org
"""
import os
import os.path

from openpyxl.reader.excel import load_workbook
from PIL import Image
from PIL.ExifTags import TAGS

images_filename = 'images.csv'
description_filename = 'description.txt'

#version 1.0 format described at https://github.com/catami/catami/wiki/Data-importing
current_format_version = '1.0'
fill_value = -999.


def get_camera_makemodel(image_name):
    """Returns a camera make and model from the exif data of an PIL Image item."""
    image = Image.open(image_name)
    info = image._getexif()
    if info:
        make = ''
        model = ''
        for tag, value in info.items():
            decoded = TAGS.get(tag, tag)
            if decoded == "Model":
                model = value
            if decoded == "Make":
                make = value
        make_model_string = make+model
    else:
        make_model_string = 'null'

    return make_model_string

if __name__ == "__main__":

    wb = load_workbook(filename=r'NingalooMar12_ImageLocations.xlsx', use_iterators=True)

    ws = wb.get_sheet_by_name(name='Sheet1')

    # AIMS xlsx format
    # row[0] -> observation file (a csv file)
    # row[1] -> path (from original file system)
    # row[2] -> Latitude (degrees)
    # row[3] -> Longitude (degrees)
    # row[4] -> Depth (m)
    # row[5] -> Date Time Original (contains date and time)
    # row[6] -> Record Created (contains date and time)
    # row[7] -> Tag (? blank in example data)
    # row[8] -> Seagrass Cover (? blank in example data)

    for row in ws.iter_rows():
        if (row[0].internal_value == 'OBSFILE'):
            continue

        obs_file_name = row[0].internal_value
        image_original_file_path = row[1].internal_value
        latitude = row[2].internal_value
        longitude = row[3].internal_value
        depth = row[4].internal_value
        image_datetime = row[5].internal_value
        record_datetime = row[6].internal_value
        tag_string = row[7].internal_value
        seagrass_cover = row[8].internal_value
        camera_name = 'null'
        camera_angle = 'Downward'
        temperature = fill_value
        salinity = fill_value
        pitch_angle = fill_value
        roll_angle = fill_value
        yaw_angle = fill_value
        altitude = fill_value

        split_path = image_original_file_path.split("\\")

        image_path = split_path[-2]
        image_name = split_path[-1]

        # get the camera from the EXIF data, if we can
        camera_name = get_camera_makemodel(os.path.join(image_path, image_name))

        # make the descriptopm file if it doesn't exist
        if not os.path.isfile(os.path.join(split_path[-2], description_filename)):
            with open(os.path.join(split_path[-2], description_filename), "w") as f:
                version_string = 'version:'+current_format_version+'\n'
                f.write(version_string)
                deployment_type_string = 'Type: TI\n'
                f.write(deployment_type_string)
                Description_string = 'Description:'+image_path+' Transects\n'
                f.write(Description_string)

        # make the images file if it doesn't exist
        if not os.path.isfile(os.path.join(split_path[-2], images_filename)):
            with open(os.path.join(split_path[-2], images_filename), "w") as f:
                version_string = 'version:'+current_format_version+'\n'
                f.write(version_string)
                headers = 'Time ,Latitude , Longitude  , Depth  , ImageName , CameraName , CameraAngle , Temperature (celcius) , Salinity (psu) , Pitch (radians) , Roll (radians) , Yaw (radians) , Altitude (metres)\n'
                f.write(headers)

        #append to csv
        with open(os.path.join(split_path[-2], images_filename), "a") as f:
            csv_string = unicode(image_datetime)+','+str(latitude)+','+str(longitude)+','+str(depth)+','+image_name+','+camera_name+','+str(camera_angle)+','+str(temperature)+','+str(salinity)+','+str(pitch_angle)+','+str(roll_angle)+','+str(yaw_angle)+','+str(altitude)+'\n'
            f.write(csv_string)
